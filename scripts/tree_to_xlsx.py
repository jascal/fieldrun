#!/usr/bin/env python3
"""tree_to_xlsx.py — dump a fieldrun `--tree` run to a multi-tab spreadsheet.

One row per tree leaf, joining the three sections of the dump:
  * clustering  → anchor (neuron/head, layer, index), exact token count, share%
  * interpret   → the decoded "specialty" tokens routed to the leaf (top-10)
  * expert tree → depth, share%, circuit count, top tokens

Tabs: `all` (every leaf) + one tab per bucket — `nil`, `X`, `lat`, and each language
present — all with the same columns. Reuses classify_tree.py for token parsing + the
(language, function) classification.

Usage:  python scripts/tree_to_xlsx.py [TREE.txt] [-o OUT.xlsx]
"""
import os, sys, re, argparse
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import classify_tree as ct
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# section line formats ------------------------------------------------------
CLUS = re.compile(r'^\s*(e\d+)\s+anchor\s+(neuron|head)\s+L(\d+)\s+#(\d+)\s+(\d+)\s+circuits\s+(\d+)\s+tokens\s+\((\d+)%\)')
INTERP = re.compile(r'^\s*(e\d+|residual)\s+(\d+)\s+tok\s+→\s+(.*)$')
TREE = re.compile(r'^(?P<ind> +)(?P<label>(?:[a-z]+\.)*e\d+)\s+(?P<pct>\d+)%\s+\(\s*(?P<circ>\d+) circ\)\s+(?P<toks>.*)$')

def fmt_toks(toks):
    nl, esc = '\n', '\\n'                                 # kept out of the f-string (3.10 forbids backslashes there)
    return '  '.join(f"{t.replace(nl, esc)!r}·{c}" for t, c in toks)

COLUMNS = [
    ('language', 10), ('label', 9), ('depth', 6), ('function', 9), ('n_circuits', 11),
    ('tokens', 9), ('share_pct', 10), ('anchor', 20), ('anchor_layer', 13),
    ('top_token', 16), ('top_token_n', 12), ('n_tok_shown', 12), ('tokens_detail', 100),
]
NUMERIC = {'depth', 'n_circuits', 'tokens', 'share_pct', 'anchor_layer', 'top_token_n', 'n_tok_shown'}
BUCKET_ORDER = ['nil', 'X', 'lat', 'en', 'de', 'es', 'fr', 'it', 'ru', 'zh', 'ja', 'ar', 'hi']

def parse(tree_path):
    clus, interp, leaves = {}, {}, []
    for line in open(tree_path, encoding='utf-8'):
        if (m := CLUS.match(line)):
            clus[m.group(1)] = dict(kind=m.group(2), layer=int(m.group(3)), idx=int(m.group(4)),
                                    circ=int(m.group(5)), tokens=int(m.group(6)), pct=int(m.group(7)))
        elif (m := INTERP.match(line)):
            if m.group(1) != 'residual':
                interp[m.group(1)] = ct.parse_tokens(m.group(3))
        elif (m := TREE.match(line)):
            toks = ct.parse_tokens(m.group('toks'))
            if toks:
                leaves.append((m.group('label'), len(m.group('ind')) // 2 - 1,
                               int(m.group('pct')), int(m.group('circ')), toks))
    recs = []
    for label, depth, pct, circ, toks in leaves:
        lang, func = ct.classify(toks)
        c = clus.get(label)                      # depth-0 experts carry anchor + exact counts
        detail = interp.get(label, toks)         # interpret has the richer top-10; else the tree's top-6
        top_t, top_n = (detail[0] if detail else ('', 0))
        recs.append({
            'language': lang, 'label': label, 'depth': depth, 'function': func, 'n_circuits': circ,
            'tokens': c['tokens'] if c else sum(n for _, n in toks),   # exact (depth-0) / Σ-shown (deeper)
            'share_pct': pct,
            'anchor': f"{c['kind']} L{c['layer']} #{c['idx']}" if c else '',
            'anchor_layer': c['layer'] if c else '',
            'top_token': top_t.replace('\n', '\\n'), 'top_token_n': top_n,
            'n_tok_shown': len(detail), 'tokens_detail': fmt_toks(detail),
        })
    recs.sort(key=lambda r: (-r['tokens'], r['depth'], r['label']))
    return recs

def write_sheet(ws, rows):
    head_font = Font(bold=True, color='FFFFFF')
    head_fill = PatternFill('solid', fgColor='305496')
    for j, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(1, j, name)
        cell.font, cell.fill = head_font, head_fill
        cell.alignment = Alignment(vertical='center')
        ws.column_dimensions[get_column_letter(j)].width = width
    for i, r in enumerate(rows, 2):
        for j, (name, _) in enumerate(COLUMNS, 1):
            v = r[name]
            ws.cell(i, j, v if (name not in NUMERIC or v != '') else '')
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"

def main():
    ap = argparse.ArgumentParser(description="Dump a fieldrun --tree run to a multi-tab .xlsx.")
    ap.add_argument('tree', nargs='?', default='tree.txt')
    ap.add_argument('-o', '--out', help='output .xlsx (default: <tree>.xlsx)')
    a = ap.parse_args()

    recs = parse(a.tree)
    if not recs:
        sys.exit(f"no leaves parsed from {a.tree} (is it a --tree dump?)")

    wb = Workbook(); wb.remove(wb.active)
    write_sheet(wb.create_sheet('all'), recs)
    present = [b for b in BUCKET_ORDER if any(r['language'] == b for r in recs)]
    for b in present:
        write_sheet(wb.create_sheet(b), [r for r in recs if r['language'] == b])

    dst = a.out or (a.tree.rsplit('.', 1)[0] + '.xlsx')
    wb.save(dst)
    print(f"wrote {dst}  —  {len(recs)} leaves across {1 + len(present)} tabs")
    print("tabs: all(" + str(len(recs)) + ")  " +
          "  ".join(f"{b}({sum(r['language'] == b for r in recs)})" for b in present))

if __name__ == '__main__':
    main()
